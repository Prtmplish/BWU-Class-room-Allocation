"""
Parse 'Class room Allotment_2026-27_ODD.xlsx' into rooms.json for the dashboard.

Each building sheet stacks room blocks:
  - a header row in column A starting with "Room No"
  - a "Days" row whose columns 2..N hold the time-slot labels
  - day rows MON..SAT whose cells hold course codes (blank = free)

Run:  python parse_rooms.py
Output: rooms.json  (consumed by index.html build step)
"""
import json
import re
import os

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "..", "Class room Allotment_2026-27_ODD.xlsx")
OUT = os.path.join(HERE, "rooms.json")

DAYS = ["MON", "TUE", "WED", "THU", "FRI", "SAT"]

# Canonical 11 period slots (start hours on a 24h clock so we can sort/range).
CANONICAL = [
    ("08:00-09:00", 8), ("09:00-10:00", 9), ("10:00-11:00", 10),
    ("11:00-12:00", 11), ("12:00-13:00", 12), ("13:00-14:00", 13),
    ("14:00-15:00", 14), ("15:00-16:00", 15), ("16:00-17:00", 16),
    ("17:00-18:00", 17), ("18:00-19:00", 18),
]
CANON_LABELS = [c[0] for c in CANONICAL]


def norm_time(raw):
    """Map a sheet time label like '8:00 - 9:00' / '12:00-1:00' to a canonical slot."""
    if not raw:
        return None
    s = str(raw).strip().replace(" ", "")
    m = re.match(r"(\d{1,2}):\d{2}-(\d{1,2}):\d{2}", s)
    if not m:
        return None
    start = int(m.group(1))
    # afternoon slots are written on a 12h clock (1:00 = 13:00)
    if start < 8:
        start += 12
    for label, hr in CANONICAL:
        if hr == start:
            return label
    return None


def clean(v):
    if v is None:
        return ""
    return " ".join(str(v).split())


def parse_room_header(text):
    """Pull a short room number/name and a type out of the header text."""
    t = clean(text)
    t = re.sub(r"^Room No\.?", "", t).strip()
    # room number: first standalone token of digits (optionally with a letter)
    num = ""
    m = re.search(r"\b(\d{2,4}[A-Za-z]?)\b", t)
    if m:
        num = m.group(1)
    # type: first parenthesised group mentioning Room / Lab
    rtype = ""
    for grp in re.findall(r"\(([^()]*)\)", t):
        if re.search(r"room|lab", grp, re.I):
            rtype = grp.strip()
            break
    return num, rtype, t


def short_building(sheet_name):
    name = clean(sheet_name)
    m = re.search(r"\(([^()]*)\)", name)
    code = m.group(1) if m else name
    base = re.sub(r"\([^()]*\)", "", name).strip()
    return base, code


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    buildings = []
    rooms = []
    rid = 0

    for sheet in wb.sheetnames:
        if "INTERNSHIP" in sheet.upper():
            continue
        ws = wb[sheet]
        bname, bcode = short_building(sheet)
        buildings.append({"name": bname, "code": bcode, "sheet": sheet})

        maxc = ws.max_column
        r = 1
        room_rows = []
        # collect header row indices
        for rr in range(1, ws.max_row + 1):
            a = clean(ws.cell(row=rr, column=1).value)
            if a.lower().startswith("room no"):
                room_rows.append(rr)
        room_rows.append(ws.max_row + 1)  # sentinel

        for i in range(len(room_rows) - 1):
            start = room_rows[i]
            end = room_rows[i + 1]
            header = ws.cell(row=start, column=1).value
            num, rtype, full = parse_room_header(header)

            # map columns -> canonical time, using the 'Days' row in this block
            col_time = {}
            for rr in range(start, end):
                a = clean(ws.cell(row=rr, column=1).value)
                if a.lower() == "days":
                    for c in range(2, maxc + 1):
                        nt = norm_time(ws.cell(row=rr, column=c).value)
                        if nt:
                            col_time[c] = nt
                    break
            if not col_time:
                # fall back to standard layout cols 2..12
                for idx, c in enumerate(range(2, 13)):
                    col_time[c] = CANON_LABELS[idx]

            # schedule[day][timelabel] = course string ("" = free)
            schedule = {d: {lbl: "" for lbl in CANON_LABELS} for d in DAYS}
            for rr in range(start, end):
                a = clean(ws.cell(row=rr, column=1).value).upper()
                if a in DAYS:
                    for c, lbl in col_time.items():
                        schedule[a][lbl] = clean(ws.cell(row=rr, column=c).value)

            rid += 1
            rooms.append({
                "id": rid,
                "building": bname,
                "buildingCode": bcode,
                "number": num,
                "type": rtype,
                "label": full,
                "schedule": schedule,
            })

    data = {
        "generated": "2026-2027 ODD",
        "days": DAYS,
        "slots": CANON_LABELS,
        "buildings": buildings,
        "rooms": rooms,
    }
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=1)
    print(f"Parsed {len(rooms)} rooms across {len(buildings)} buildings -> {OUT}")

    # Build the self-contained dashboard by injecting the data into the template.
    tpl_path = os.path.join(HERE, "template.html")
    out_html = os.path.join(HERE, "dashboard.html")
    if os.path.exists(tpl_path):
        with open(tpl_path, encoding="utf-8") as f:
            tpl = f.read()
        # embed as a JSON string inside a <script type="application/json"> block
        payload = json.dumps(data, ensure_ascii=False)
        html = tpl.replace("/*__DATA__*/", payload)
        # inline SheetJS so the dashboard is a single self-contained file
        lib = os.path.join(HERE, "lib", "xlsx.full.min.js")
        if os.path.exists(lib):
            with open(lib, encoding="utf-8") as f:
                libjs = f.read()
            html = html.replace(
                '<script src="lib/xlsx.full.min.js"></script>',
                "<script>" + libjs + "</script>",
            )
        with open(out_html, "w", encoding="utf-8") as f:
            f.write(html)
        print(f"Built standalone dashboard -> {out_html}")


if __name__ == "__main__":
    main()
